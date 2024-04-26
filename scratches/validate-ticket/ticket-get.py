# coding: utf-8
########################################################################################################################
###########################################                       ######################################################
###########################################  READ ONLY LIBRARIES  ######################################################
###########################################                       ######################################################
########################################################################################################################
import json, logging, os, sys
sys.path.append(os.path.dirname(__file__))
log = logging.getLogger(__name__)
from datetime import datetime, timedelta
from airflow import DAG
from airflow.operators import MATInstanceInitOperator, MATInstanceExitOperator, MATPythonOperator, DummyOperator
from airflow.utils.trigger_rule import TriggerRule
from airflow.exceptions import AirflowException
from mat import *
from mat_error_management import MATWorkflowErrorManagement
from mat_success_management import MATWorkflowSuccessManagement
from mat_runtime import *
########################################################################################################################
#######################################                                  ###############################################
#######################################  CUSTOM LIBRARIES AND VARIABLES  ###############################################
#######################################                                  ###############################################
########################################################################################################################

try:
    import usecase
except:
    pass
import xmltodict

########################################################################################################################
##############################################                  ########################################################
##############################################  DAG DEFINITION  ########################################################
##############################################                  ########################################################
########################################################################################################################


default_args = {
    'owner': 'Iquall',
    'depends_on_past': False,
    'provide_context': True
}

dag = DAG(dag_id='7wl-9r5-czg', description='', start_date=datetime(2021,12,1), schedule_interval=None, catchup=False, on_failure_callback=MATWorkflowErrorManagement, on_success_callback=MATWorkflowSuccessManagement, default_args=default_args)


########################################################################################################################
##############################################                    ######################################################
##############################################  CUSTOM FUNCTIONS  ######################################################
##############################################                    ######################################################
########################################################################################################################


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Provide the path to your original Excel file
input_excel_path = r"C:\Users\aw474y\OneDrive - AT&T Mexico\Escritorio\rutina\4.xlsx"

# Convert the Excel file to a CSV file
csv_path = r"C:\Users\aw474y\OneDrive - AT&T Mexico\Escritorio\rutina\Huawei-4G.csv"
df = pd.read_excel(input_excel_path)
df.to_csv(csv_path, index=False)

# Group by the value in the 'C2' column and write to individual Excel files
grouped = df.groupby('C2')
for name, group in grouped:
    filename = os.path.join("C:\\Users\\aw474y\\OneDrive - AT&T Mexico\\Escritorio\\rutina\\", f"{name}.xlsx")
    group.to_excel(filename, index=False)

    # Load the workbook
    book = load_workbook(filename)

    # Now you can access the sheet you want to work with
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    sheet = writer.sheets['Sheet1']

    # Define cell color formats
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    blue_fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")
    green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

    # Color cells based on the conditions defined in "ProcessWorksheet" of the original VBA
    for row in sheet.iter_rows(min_row=2):  # start from 2nd row to skip the header
        for cell in row:
            if isinstance(cell.value, str) and "Target_String" in cell.value:
                cell.fill = red_fill
            elif isinstance(cell.value, float) and cell.value > 1.0:
                # Change the condition as per your requirement
                cell.fill = blue_fill

    writer.save()


def get_tickets(user, pwsd):
    url = 'https://your_api_endpoint'
    response = requests.get(url, auth=(user, pwsd))
    data = response.json()
    return data

 class TicketValidator:
     # The rest of your class...

     def validate_gestor_field(self, gestor_field_value):
         # Replace this with your specific validation logic for 'gestor' field
         return bool(gestor_field_value)

     def validate_NOC_field(self, NOC_field_value):
         # Replace this with your specific validation logic for 'NOC' field
         return bool(NOC_field_value)

     def validate_ticket(self, ticket):
         # Check if IP is duplicated
         if self.check_ip_duplicate(ticket['ip']):
             print(f"Validation failed for the ticket with ID {ticket['id']}: IP is duplicated.")
             return False

         # Check 'gestor' field
         if 'gestor' not in ticket or not self.validate_gestor_field(ticket['gestor']):
             print(f"Validation failed for the ticket with ID {ticket['id']}: Invalid 'gestor' field.")
             return False

         # Check 'NOC' field
         if 'NOC' not in ticket or not self.validate_NOC_field(ticket['NOC']):
             print(f"Validation failed for the ticket with ID {ticket['id']}: Invalid 'NOC' field.")
             return False

         return True

     def get_tickets_assigned_to(user, tickets):
         # assuming tickets is a list of dictionary objects
         return [ticket for ticket in tickets if ticket['assigned_to'] == user]

     # get all tickets
     all_tickets = get_all_tickets_from_redemy()

     # get tickets assigned to 'BO performance'
     bo_performance_tickets = get_tickets_assigned_to('BO performance', all_tickets)

     # get the last few tickets assigned to 'BO performance'
     last_few_tickets = bo_performance_tickets[-5:]


     def validate_ticket_kpi(ticket, kpi, statistic):
         # Assuming 'kpi_values' is a dict containing various KPIs for the ticket
         ticket_kpi = ticket['kpi_values'][kpi]

         if ticket_kpi > statistic:
             return True
         else:
             return False
        # Assuming 'recent_tickets' is a list of your recent tickets
        for ticket in recent_tickets:
            if validate_ticket_kpi(ticket, kpi='response_time', statistic=average_response_time):
                # The response time of this ticket is above average
                pass  # Replace with what you want to do
            else:
                # The response time of this ticket is not above average
                pass  # Replace with what you want to do

